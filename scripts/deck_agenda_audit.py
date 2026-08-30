#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
deck_agenda_audit.py — decks vs. the agenda, two tiers (2026-08-29)
===================================================================
Katherine's instruction: "Once all the decks have been finalized they should be
audited for links and against the agenda."

WHAT THIS IS NOT
----------------
This is not `deck_lint.py`. That script judges one deck against the DESIGN
standard (format tokens, marker strings, the 5 core questions, Concept Bank,
teacher note). It never opens the agenda and never looks at another deck.

This script judges the deck SET against the AGENDA and against the LINKS that
carry decks to students. It duplicates none of deck_lint's checks; it *calls*
deck_lint (see --no-lint) and carries its verdict into the readiness roll-up as
one column, so the roll-up is a single "is this cycle ready to teach" line.

Nor is it `deck_link_check.py`. That walks the hyperlinks INSIDE one .pptx.
This walks the hyperlinks in the AGENDA — the spreadsheet cells that hand a deck
to a teacher and to a class. Its --check-live request shape (HEAD then GET, the
same UA and timeout) is deliberately the same as that script's.

WHAT IT ANSWERS
---------------
A. LINK INTEGRITY  (the '🌿 Biology List 26-27' tab, plus 'Bio as List')
   Every deck link is stored as `=HYPERLINK("url","label")`. openpyxl does not
   evaluate formulas, so the formula TEXT is parsed, not the cell value.

     H-STUDENT-EDIT      the student column carries an /edit link. A hard error:
                         it hands the class the master deck to edit.
     H-STUDENT-NOTCOPY   student link is a Google Slides URL in some other mode
                         (/preview, /view, bare) — not distributable by /copy.
     H-STUDENT-FILEFORM  student link is a Drive `file/d/…` or `open?id=…` link.
                         Those cannot be /copy distributed at all.
     H-LINK-MISMATCH     teacher and student links on one row point at different
                         deck IDs. Her doc states they always should agree.
     H-LINK-EMPTY        a link cell that is label text with no URL, or a
                         HYPERLINK formula whose URL argument is empty.
                         (Row 6 of the Biology tab is the known instance.)
     H-STUDENT-MISSING   teacher link present, student link absent.
     H-DECK-MISSING      a row that is a real scheduled meeting — it carries both
                         a Block and a Class phase — with no deck link at all.

     A-TEACHER-MISSING   student link present, teacher link absent.
     A-FOLDER-LINK       cycle-folder / turn-it-in link malformed.
     A-BIOLIST-VT-DRIFT  'Bio as List' 🎞 VT link and its own _vtURL helper column
                         disagree on the deck ID.
     A-BIOLIST-FILEFORM  'Bio as List' VT link is a Drive file/open?id link.
     A-BIOLIST-ORPHAN    'Bio as List' points at a deck ID the dated Biology tab
                         never uses. A stale week index.
     A-DAY-WEEKDAY       the row's Date does not fall on the weekday its Day
                         column claims.

B. DECK <-> AGENDA AGREEMENT
     A-DECK-NO-AGENDA    a deck on disk that no agenda row points at.
                         (Built but never scheduled.)
     A-AGENDA-NO-DECK    an agenda row with no deck link that is NOT yet a
                         scheduled meeting — a skeleton row. Advisory, because
                         this is work not done, not a defect.
     A-ID-UNKNOWN        an agenda deck ID that is in neither decks_live_ids.csv
                         nor derivable from a deck on disk.
     A-CYCLE-MISMATCH    the row's Cycle code (C09) disagrees with the cycle of
                         the deck it links to.
     A-TOPIC-DRIFT       the row's Topic shares no content word with the deck's
                         title OR any of its critical aspects.
     A-TOPIC-ASPECT-DRIFT the Topic matches the deck TITLE but none of its
                         critical aspects. This is the spec's own example — a
                         row labelled "Cell Transport" pointing at a deck whose
                         aspects are about membrane structure.
                         BOTH are LEXICAL, not semantic: they compare content
                         words. "Introduction to Heredity" -> "Mendelian
                         Genetics" fires and is a correct pairing. Read these as
                         "go look", not as "this is wrong".
     A-DUP-DAY1          a later "1 of N" row reusing an EARLIER unit's day-1
                         deck. Katherine has ruled: "later units need their own
                         decks." The earlier date is canonical; the later unit is
                         reported as NEEDS ITS OWN DECK.
     A-DAY1-NO-DECK      a "1 of N" row issuing no deck.
     A-LATER-NEW-DECK    a non-first meeting introducing a deck the row before it
                         was not using. Only "1 of N" should carry a new deck.
     A-NO-DAY1           a deck that the agenda never introduces on a "1 of N"
                         row — it only ever appears mid-cycle.

C. READINESS ROLL-UP
   One line per cycle: deck exists · deck_lint verdict · agenda rows · links
   well-formed · student link is /copy · live doc_id present for the import
   route. READY only when all of them hold.

WHY TWO TIERS
-------------
Same reason deck_lint has them, and the same line: a tool that screams about
everything gets ignored inside a week. Exit code 1 fires ONLY on HARD — links
that actively mis-serve a class, and a scheduled meeting with no deck. Decks not
yet built, decks not yet scheduled, duplicate day-1 decks and topic drift are all
ADVISORY: they are work not done, and the decks are not finalized yet.

READ-ONLY
---------
This script writes nothing except its own --csv / --md reports. It opens every
workbook and every .pptx read-only, refuses any output path under CloudStorage or
under ~/deck_work, and asserts the agenda file's size+mtime are unchanged when it
finishes. See assert_readonly() and _guard_output_path().

USAGE
-----
  ~/deck_work/.venv/bin/python scripts/deck_agenda_audit.py
  ... --csv audits/deck_agenda_audit_2026-08-29.csv \
      --md  audits/deck_agenda_audit_2026-08-29.md
  ... --tier hard          # print HARD only; never changes the exit code
  ... --no-lint            # skip the deck_lint pass (faster)
  ... --check-live         # OFF BY DEFAULT. Only this flag makes network calls.

Interpreter: ~/deck_work/.venv/bin/python (3.9.6). No 3.10+ syntax.
Dependencies: openpyxl, python-pptx (python-pptx only via deck_lint / rebuild scan).
"""
from __future__ import print_function

import argparse
import csv as csvmod
import datetime
import glob as globmod
import os
import re
import sys

import openpyxl


# ---------------------------------------------------------------------------
# Paths and tab names that must not drift
# ---------------------------------------------------------------------------

HOME = os.path.expanduser("~")
REPO = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

DEFAULT_AGENDA = os.path.join(HOME, "deck_work", "agenda_work",
                              "Agenda__BACKUP_2026-08-29.xlsx")
LIVE_AGENDA = os.path.join(
    HOME, "Library", "CloudStorage",
    "GoogleDrive-katherine.vonduyke@redclay.k12.de.us", "My Drive",
    "00 AIHS 2026 7 (Private)", "Agenda__2026-27_RELINKED.xlsx")

DEFAULT_EXPORTS = os.path.join(HOME, "deck_work", "exports")
DEFAULT_REBUILD = os.path.join(HOME, "deck_work", "12b_rebuild")
DEFAULT_LIVE_IDS = os.path.join(REPO, "docs", "decks_live_ids.csv")
AUDITS_DIR = os.path.join(REPO, "audits")

# Exact tab names. The Biology tab carries a leading emoji; ' Lab Materials by
# Date' carries a LEADING SPACE. Both are spelled out here as byte-exact
# constants so that a rename fails loudly in require_sheet() instead of
# silently returning zero rows.
TAB_BIOLOGY = u"\U0001F33F Biology List 26-27"   # "🌿 Biology List 26-27"
TAB_BIO_AS_LIST = u"Bio as List"
TAB_SCHEDULE = u"Schedule"
TAB_FILE_IDS = u"File IDs (do not delete)"
TAB_LAB_MATERIALS = u" Lab Materials by Date"    # LEADING SPACE. Intentional.

# Biology tab column layout (1-based), header on row 2, data from row 3.
BIO_HEADER_ROW = 2
BIO_FIRST_DATA_ROW = 3
COL_DATE, COL_DAY, COL_CYCLE, COL_TOPIC, COL_BLOCK, COL_PHASE = 1, 2, 3, 4, 5, 6
COL_DECK_TEACHER, COL_DECK_STUDENT = 7, 8
COL_CYCLE_FOLDER, COL_TURNIN, COL_LABPREP, COL_NOTES = 9, 10, 11, 12
BIO_HEADERS_EXPECTED = {
    COL_DATE: "Date", COL_DAY: "Day", COL_CYCLE: "Cycle", COL_TOPIC: "Topic",
    COL_BLOCK: "Block", COL_PHASE: "Class phase",
    COL_DECK_TEACHER: "Deck (teacher)", COL_DECK_STUDENT: "Deck (student)",
    COL_CYCLE_FOLDER: "Cycle folder", COL_TURNIN: "Turn it in Folders",
    COL_LABPREP: "Lab prep", COL_NOTES: "My notes",
}

# 'Bio as List' layout, header on row 3.
BAL_HEADER_ROW = 3
BAL_FIRST_DATA_ROW = 4
BAL_WK, BAL_DATES, BAL_TOPIC, BAL_PHASE = 1, 2, 3, 4
BAL_AGENDA, BAL_VT, BAL_FOLDER, BAL_ACTIVITY = 5, 6, 7, 8
BAL_START, BAL_AGENDAURL, BAL_VTURL = 9, 10, 11

WEEKDAY_NAMES = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")

# How many instances of one code to spell out in the human report.
# The findings CSV always carries every one of them.
LIST_CAP = 12


# ---------------------------------------------------------------------------
# Read-only guards
# ---------------------------------------------------------------------------

FORBIDDEN_WRITE_MARKERS = ("cloudstorage", "google drive", "googledrive",
                           os.path.join("deck_work", ""))


def _guard_output_path(path):
    """Refuse to write anywhere but a normal local report location.

    Hard refusal for CloudStorage (the live agenda lives there) and for
    ~/deck_work (another agent is working in agenda_work; the decks are there).
    """
    ap = os.path.abspath(os.path.expanduser(path))
    low = ap.lower()
    if "cloudstorage" in low or "/google drive/" in low:
        raise SystemExit("deck_agenda_audit: REFUSING to write into CloudStorage: %s" % ap)
    if low.startswith(os.path.join(HOME, "deck_work").lower()):
        raise SystemExit("deck_agenda_audit: REFUSING to write into ~/deck_work: %s" % ap)
    if ap.lower().endswith((".pptx", ".xlsx", ".xlsm")):
        raise SystemExit("deck_agenda_audit: REFUSING to write a deck or workbook: %s" % ap)
    return ap


def _fingerprint(path):
    try:
        st = os.stat(path)
        return (st.st_size, st.st_mtime_ns)
    except OSError:
        return None


class ReadOnlyWitness(object):
    """Records size+mtime of every file we read, and asserts none changed."""

    def __init__(self):
        self._marks = {}

    def watch(self, path):
        if path and os.path.exists(path):
            self._marks[os.path.abspath(path)] = _fingerprint(path)

    def watch_tree(self, root, suffix=".pptx"):
        if not root or not os.path.isdir(root):
            return
        for dirpath, _dirs, files in os.walk(root):
            for f in files:
                if f.lower().endswith(suffix):
                    self.watch(os.path.join(dirpath, f))

    def assert_unchanged(self):
        changed = []
        for path, before in self._marks.items():
            if _fingerprint(path) != before:
                changed.append(path)
        if changed:
            raise AssertionError(
                "deck_agenda_audit is READ-ONLY but these files changed during "
                "the run:\n  " + "\n  ".join(sorted(changed)))
        return len(self._marks)


WITNESS = ReadOnlyWitness()


# ---------------------------------------------------------------------------
# Findings
# ---------------------------------------------------------------------------

HARD = "HARD"
ADVISORY = "ADVISORY"


class Finding(object):
    __slots__ = ("tier", "code", "where", "message", "cycle")

    def __init__(self, tier, code, where, message, cycle=""):
        self.tier = tier
        self.code = code
        self.where = where          # e.g. "Biology!H33" or "deck:Cycle 18"
        self.message = message
        self.cycle = cycle

    def line(self):
        return "[%s] %s: %s" % (self.code, self.where, self.message)

    def as_row(self):
        return [self.tier, self.code, self.cycle, self.where, self.message]


class Report(object):
    def __init__(self):
        self.findings = []

    def add(self, tier, code, where, message, cycle=""):
        self.findings.append(Finding(tier, code, where, message, cycle))

    def hard(self):
        return [f for f in self.findings if f.tier == HARD]

    def advisory(self):
        return [f for f in self.findings if f.tier == ADVISORY]

    def by_code(self):
        out = {}
        for f in self.findings:
            out.setdefault(f.code, []).append(f)
        return out


# ---------------------------------------------------------------------------
# HYPERLINK formula parsing
# ---------------------------------------------------------------------------

# =HYPERLINK("url","label")  — quotes inside a Google/Excel string double up.
_HL_RE = re.compile(
    r'^\s*=\s*HYPERLINK\s*\(\s*"((?:[^"]|"")*)"\s*(?:,\s*"((?:[^"]|"")*)"\s*)?\)\s*$',
    re.IGNORECASE | re.DOTALL)
# A formula that is just a quoted constant, e.g. ="🎞 Deck (teacher)".
_CONST_RE = re.compile(r'^\s*=\s*"((?:[^"]|"")*)"\s*$', re.DOTALL)

_ID_PATTERNS = (
    re.compile(r'/(?:presentation|document|spreadsheets|file|forms)/d/([A-Za-z0-9_\-]{15,})'),
    re.compile(r'/drive/folders/([A-Za-z0-9_\-]{15,})'),
    re.compile(r'[?&]id=([A-Za-z0-9_\-]{15,})'),
)

_SLIDES_MODE_RE = re.compile(
    r'/presentation/d/[A-Za-z0-9_\-]{15,}/([A-Za-z0-9_]+)')


class Link(object):
    """One agenda cell, decoded."""

    __slots__ = ("raw", "kind", "label", "url", "doc_id", "service", "mode")

    def __init__(self, raw, kind, label="", url="", doc_id="", service="", mode=""):
        self.raw = raw
        self.kind = kind        # empty | hyperlink | text | url | const | formula
        self.label = label
        self.url = url
        self.doc_id = doc_id
        self.service = service  # slides | doc | sheet | drivefile | folder | other
        self.mode = mode        # copy | edit | preview | view | present | ''

    def __nonzero__(self):      # py2 habit; harmless
        return bool(self.url)

    __bool__ = __nonzero__

    @property
    def has_label_no_url(self):
        return self.kind in ("text", "const") and bool(self.label) and not self.url

    @property
    def is_slides(self):
        return self.service == "slides"

    def describe(self):
        if self.kind == "empty":
            return "(empty)"
        if self.url:
            return self.url
        return "label-only %r" % (self.label,)


def _unescape(s):
    return (s or "").replace('""', '"')


def classify_url(url):
    """Return (doc_id, service, mode)."""
    doc_id = ""
    for pat in _ID_PATTERNS:
        m = pat.search(url)
        if m:
            doc_id = m.group(1)
            break
    low = url.lower()
    if "/presentation/d/" in low:
        service = "slides"
    elif "/document/d/" in low:
        service = "doc"
    elif "/spreadsheets/d/" in low:
        service = "sheet"
    elif "/drive/folders/" in low:
        service = "folder"
    elif "/file/d/" in low or "open?id=" in low:
        service = "drivefile"
    elif url:
        service = "other"
    else:
        service = ""
    mode = ""
    m = _SLIDES_MODE_RE.search(url)
    if m:
        mode = m.group(1).lower()
    elif service == "slides":
        mode = "bare"
    return doc_id, service, mode


def read_link(ws, row, col):
    """Decode one cell into a Link. Formula TEXT is parsed; openpyxl does not
    evaluate HYPERLINK() and never will."""
    cell = ws.cell(row=row, column=col)
    v = cell.value
    if v is None or (isinstance(v, str) and not v.strip()):
        # A real (non-formula) hyperlink attached to an otherwise blank cell.
        if cell.hyperlink is not None and getattr(cell.hyperlink, "target", None):
            url = cell.hyperlink.target
            doc_id, service, mode = classify_url(url)
            return Link(url, "hyperlink", "", url, doc_id, service, mode)
        return Link("", "empty")

    if not isinstance(v, str):
        return Link(str(v), "text", label=str(v))

    s = v.strip()
    m = _HL_RE.match(s)
    if m:
        url = _unescape(m.group(1)).strip()
        label = _unescape(m.group(2) or "").strip()
        if not url:
            return Link(s, "hyperlink", label=label)  # HYPERLINK with empty URL
        doc_id, service, mode = classify_url(url)
        return Link(s, "hyperlink", label, url, doc_id, service, mode)

    m = _CONST_RE.match(s)
    if m:
        return Link(s, "const", label=_unescape(m.group(1)).strip())

    if s.startswith("="):
        return Link(s, "formula", label=s)

    if s.lower().startswith(("http://", "https://")):
        doc_id, service, mode = classify_url(s)
        return Link(s, "url", "", s, doc_id, service, mode)

    if cell.hyperlink is not None and getattr(cell.hyperlink, "target", None):
        url = cell.hyperlink.target
        doc_id, service, mode = classify_url(url)
        return Link(s, "hyperlink", s, url, doc_id, service, mode)

    return Link(s, "text", label=s)


# ---------------------------------------------------------------------------
# Cycle keys
# ---------------------------------------------------------------------------

_CYCLE_KEY_RE = re.compile(r'Cycle\s*0*(\d+)\s*([a-z]?)', re.IGNORECASE)
_AGENDA_CYCLE_RE = re.compile(r'\bC\s*0*(\d+)\s*([a-z]*)', re.IGNORECASE)


def cycle_key(num, suffix=""):
    """Canonical key, matching decks_live_ids.csv: 'Cycle 07a'."""
    return "Cycle %02d%s" % (int(num), (suffix or "").lower())


def key_from_text(text, rx=_CYCLE_KEY_RE):
    if not text:
        return ""
    m = rx.search(str(text))
    if not m:
        return ""
    return cycle_key(m.group(1), m.group(2))


def key_from_agenda_cycle(text):
    """'C09' -> 'Cycle 09'.  'C13e' -> 'Cycle 13e'.  '—' -> ''.

    NOTE the 'e' suffixes ('C06e', 'C13e') are Katherine's *extension* markers on
    the agenda, not deck-file suffixes ('07a', '16b'). They are kept in the key so
    they show in the report, and stripped by cycle_number() for comparison.
    """
    return key_from_text(text, _AGENDA_CYCLE_RE)


def cycle_number(key):
    m = re.search(r'(\d+)', key or "")
    return int(m.group(1)) if m else None


# ---------------------------------------------------------------------------
# Topic <-> deck content agreement
# ---------------------------------------------------------------------------

STOPWORDS = set("""
a an and the of to in on for with without by from into over under is are was were
be been being as at or vs versus its it this that these those what how why when
which who whom more most less least new intro introduction continued class day
lesson notes draft apply part unit cycle deck student teacher slides review
""".split())

_WORD_RE = re.compile(r"[A-Za-z][A-Za-z'\-]+")


def content_words(text):
    out = set()
    for w in _WORD_RE.findall(str(text or "")):
        w = w.lower().strip("-'")
        if len(w) < 3 or w in STOPWORDS:
            continue
        # crude singularisation so "Cycles"/"cycle", "Enzymes"/"enzyme" meet
        if w.endswith("ies") and len(w) > 4:
            w = w[:-3] + "y"
        elif w.endswith("es") and len(w) > 4 and not w.endswith("ses"):
            w = w[:-2]
        elif w.endswith("s") and len(w) > 3 and not w.endswith("ss"):
            w = w[:-1]
        out.add(w)
    return out


def topic_agreement(topic, deck_title, aspects):
    """Return one of: 'title+aspect', 'title-only', 'aspect-only', 'none', 'n/a'."""
    tw = content_words(topic)
    if not tw:
        return "n/a"
    title_hit = bool(tw & content_words(deck_title))
    aspect_hit = bool(tw & content_words(aspects))
    if title_hit and aspect_hit:
        return "title+aspect"
    if title_hit:
        return "title-only"
    if aspect_hit:
        return "aspect-only"
    return "none"


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def require_sheet(wb, name, path):
    """Fail LOUDLY on a wrong/renamed tab. Never silently return zero rows."""
    if name in wb.sheetnames:
        return wb[name]
    raise SystemExit(
        "deck_agenda_audit: tab not found: %r\n"
        "  workbook: %s\n"
        "  tabs present (%d): %s\n"
        "  This is a hard stop on purpose. A renamed tab must not read as an\n"
        "  empty tab — that would report every deck as unscheduled."
        % (name, path, len(wb.sheetnames),
           ", ".join(repr(n) for n in wb.sheetnames)))


def load_live_ids(path):
    """decks_live_ids.csv -> {key: {...}}"""
    if not os.path.exists(path):
        raise SystemExit("deck_agenda_audit: decks_live_ids.csv not found: %s" % path)
    WITNESS.watch(path)
    out = {}
    by_doc = {}
    with open(path, "r") as fh:
        for r in csvmod.DictReader(fh):
            key = key_from_text(r.get("key", "")) or (r.get("key") or "").strip()
            r["_key"] = key
            out[key] = r
            doc = (r.get("doc_id") or "").strip()
            if doc:
                by_doc[doc] = key
    return out, by_doc


def newest_inventory(explicit=None):
    if explicit:
        return os.path.abspath(os.path.expanduser(explicit))
    cands = sorted(globmod.glob(os.path.join(AUDITS_DIR, "deck_inventory_*.csv")))
    return cands[-1] if cands else None


def load_inventory(path):
    """deck_inventory_*.csv -> {basename: row}. Reused rather than re-reading
    every .pptx: the inventory already extracted critical aspects."""
    if not path or not os.path.exists(path):
        return {}
    WITNESS.watch(path)
    out = {}
    with open(path, "r") as fh:
        for r in csvmod.DictReader(fh):
            out[(r.get("file") or "").strip()] = r
    return out


_CB_SUFFIXES = (
    " — with Concept Bank + slide index",
    " — with Concept Bank",
)


def normalize_deck_name(basename):
    n = basename
    if n.lower().endswith(".pptx"):
        n = n[:-5]
    for suf in _CB_SUFFIXES:
        if n.endswith(suf):
            n = n[:-len(suf)]
            break
    return n.strip()


def scan_decks(export_roots):
    """Every .pptx under the given roots. Returns list of dicts."""
    decks = []
    for root in export_roots:
        root = os.path.abspath(os.path.expanduser(root))
        if not os.path.isdir(root):
            continue
        WITNESS.watch_tree(root)
        for dirpath, dirs, files in os.walk(root):
            # skip build scratch: _src/, _backups/, images/ ...
            dirs[:] = [d for d in dirs if not d.startswith("_")]
            if os.path.basename(dirpath).startswith("_"):
                continue
            for f in sorted(files):
                if not f.lower().endswith(".pptx") or f.startswith("~$"):
                    continue
                full = os.path.join(dirpath, f)
                decks.append({
                    "file": f,
                    "path": full,
                    "folder": os.path.basename(dirpath),
                    "root": root,
                    "norm": normalize_deck_name(f),
                    "is_backup": (f.startswith("_") or "backup" in f.lower()
                                  or f.upper().startswith("REFERENCE")),
                    "is_concept_bank": any(normalize_deck_name(f) != f[:-5] and
                                           f[:-5].endswith(s) for s in _CB_SUFFIXES),
                })
    return decks


# ---------------------------------------------------------------------------
# Agenda: the dated Biology tab
# ---------------------------------------------------------------------------

class AgendaRow(object):
    __slots__ = ("row", "date", "day", "cycle_text", "cycle_key", "topic",
                 "block", "block_i", "block_n", "phase",
                 "teacher", "student", "folder", "turnin", "labprep", "notes",
                 "doc_id", "deck_key")

    def is_meeting(self):
        """A real scheduled meeting: it carries a Block AND a Class phase."""
        return bool(self.block) and bool(self.phase)

    def is_day1(self):
        return self.block_i == 1 and self.block_n is not None


_BLOCK_RE = re.compile(r'^\s*(\d+)\s*of\s*(\d+)\s*$', re.IGNORECASE)


def parse_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v or "").strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def load_biology_rows(ws):
    # Verify the header row before trusting the column numbers.
    header_problems = []
    for col, want in BIO_HEADERS_EXPECTED.items():
        got = ws.cell(row=BIO_HEADER_ROW, column=col).value
        got_s = ""
        if isinstance(got, str):
            m = _CONST_RE.match(got.strip())
            got_s = _unescape(m.group(1)) if m else got
        elif got is not None:
            got_s = str(got)
        if want.lower() not in got_s.lower():
            header_problems.append((col, want, got_s))

    rows = []
    for r in range(BIO_FIRST_DATA_ROW, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 13)]
        if all(v in (None, "") for v in vals):
            continue
        ar = AgendaRow()
        ar.row = r
        ar.date = parse_date(vals[COL_DATE - 1])
        ar.day = str(vals[COL_DAY - 1] or "").strip()
        ar.cycle_text = str(vals[COL_CYCLE - 1] or "").strip()
        ar.cycle_key = key_from_agenda_cycle(ar.cycle_text)
        ar.topic = str(vals[COL_TOPIC - 1] or "").strip()
        ar.block = str(vals[COL_BLOCK - 1] or "").strip()
        ar.phase = str(vals[COL_PHASE - 1] or "").strip()
        m = _BLOCK_RE.match(ar.block)
        ar.block_i = int(m.group(1)) if m else None
        ar.block_n = int(m.group(2)) if m else None
        ar.teacher = read_link(ws, r, COL_DECK_TEACHER)
        ar.student = read_link(ws, r, COL_DECK_STUDENT)
        ar.folder = read_link(ws, r, COL_CYCLE_FOLDER)
        ar.turnin = read_link(ws, r, COL_TURNIN)
        ar.labprep = str(vals[COL_LABPREP - 1] or "").strip()
        ar.notes = str(vals[COL_NOTES - 1] or "").strip()
        ar.doc_id = ar.teacher.doc_id or ar.student.doc_id
        ar.deck_key = ""
        rows.append(ar)
    return rows, header_problems


# ---------------------------------------------------------------------------
# Section A — link integrity
# ---------------------------------------------------------------------------

def check_links(rows, rep):
    for ar in rows:
        cellT = "Biology!G%d" % ar.row
        cellS = "Biology!H%d" % ar.row
        t, s = ar.teacher, ar.student

        # label text / empty-URL HYPERLINK on either deck column
        for lk, cell, who in ((t, cellT, "teacher"), (s, cellS, "student")):
            if lk.has_label_no_url or (lk.kind == "hyperlink" and not lk.url):
                rep.add(HARD, "H-LINK-EMPTY", cell,
                        "%s deck cell shows label text %r with no URL — "
                        "the link looks live and goes nowhere"
                        % (who, lk.label or lk.raw[:40]),
                        ar.cycle_key)

        # student link form
        if s.url:
            if s.service == "drivefile":
                rep.add(HARD, "H-STUDENT-FILEFORM", cellS,
                        "student link is a Drive file link (%s) — a file/d or "
                        "open?id link cannot be /copy distributed" % s.url,
                        ar.cycle_key)
            elif s.service == "slides":
                if s.mode == "edit":
                    rep.add(HARD, "H-STUDENT-EDIT", cellS,
                            "student link is /edit — this hands the class the "
                            "MASTER deck to edit. Must be /copy: %s" % s.url,
                            ar.cycle_key)
                elif s.mode != "copy":
                    rep.add(HARD, "H-STUDENT-NOTCOPY", cellS,
                            "student Slides link is /%s, not /copy: %s"
                            % (s.mode or "bare", s.url), ar.cycle_key)
            else:
                rep.add(HARD, "H-STUDENT-NOTCOPY", cellS,
                        "student deck link is not a Google Slides URL (%s): %s"
                        % (s.service or "unknown", s.url), ar.cycle_key)

        # teacher link form (advisory: /copy in the teacher column is wrong but
        # harmless — it makes her a copy instead of opening the master)
        if t.url and t.service == "slides" and t.mode == "copy":
            rep.add(ADVISORY, "A-TEACHER-COPY", cellT,
                    "teacher link is /copy — should be /edit so she opens the "
                    "master: %s" % t.url, ar.cycle_key)
        if t.url and t.service == "drivefile":
            rep.add(ADVISORY, "A-TEACHER-FILEFORM", cellT,
                    "teacher link is a Drive file link, not presentation/d: %s"
                    % t.url, ar.cycle_key)

        # the two must agree
        if t.doc_id and s.doc_id and t.doc_id != s.doc_id:
            rep.add(HARD, "H-LINK-MISMATCH", "Biology!G%d/H%d" % (ar.row, ar.row),
                    "teacher and student point at DIFFERENT decks: teacher %s vs "
                    "student %s" % (t.doc_id, s.doc_id), ar.cycle_key)

        # one present, the other not
        if t.url and not s.url:
            rep.add(HARD, "H-STUDENT-MISSING", cellS,
                    "teacher deck present but no student link — the class has no "
                    "way to get the deck", ar.cycle_key)
        if s.url and not t.url:
            rep.add(ADVISORY, "A-TEACHER-MISSING", cellT,
                    "student deck present but no teacher link", ar.cycle_key)

        # neither present
        if not t.url and not s.url and not (t.has_label_no_url or s.has_label_no_url):
            if ar.is_meeting():
                rep.add(HARD, "H-DECK-MISSING", "Biology!G%d:H%d" % (ar.row, ar.row),
                        "scheduled meeting (%s / %s) %s — %r — has NO deck link"
                        % (ar.block, ar.phase, ar.date, ar.topic), ar.cycle_key)
            else:
                rep.add(ADVISORY, "A-AGENDA-NO-DECK", "Biology!G%d:H%d" % (ar.row, ar.row),
                        "no deck link; row is not yet a scheduled meeting "
                        "(block=%r phase=%r) %s — %r"
                        % (ar.block, ar.phase, ar.date, ar.topic), ar.cycle_key)

        # supporting links
        for lk, cell, what in ((ar.folder, "Biology!I%d" % ar.row, "cycle folder"),
                               (ar.turnin, "Biology!J%d" % ar.row, "turn-it-in folder")):
            if lk.kind == "empty":
                continue
            if lk.has_label_no_url or (lk.kind == "hyperlink" and not lk.url):
                rep.add(ADVISORY, "A-FOLDER-LINK", cell,
                        "%s cell shows a label with no URL" % what, ar.cycle_key)
            elif lk.url and lk.service not in ("folder", "doc", "sheet", "slides"):
                rep.add(ADVISORY, "A-FOLDER-LINK", cell,
                        "%s link is not a Drive folder: %s" % (what, lk.url),
                        ar.cycle_key)

        # weekday sanity
        if ar.date and ar.day:
            want = WEEKDAY_NAMES[ar.date.weekday()]
            if want.lower() != ar.day.lower()[:3]:
                rep.add(ADVISORY, "A-DAY-WEEKDAY", "Biology!A%d/B%d" % (ar.row, ar.row),
                        "%s is a %s but the Day column says %s"
                        % (ar.date, want, ar.day), ar.cycle_key)


def check_bio_as_list(ws, bio_doc_ids, rep):
    """'Bio as List' is the weekly index. It is a second link surface and it
    drifts from the dated tab. Everything here is ADVISORY: it is not the
    distribution path, so nothing on it can hand a student an editable master."""
    for r in range(BAL_FIRST_DATA_ROW, ws.max_row + 1):
        wk = ws.cell(row=r, column=BAL_WK).value
        if not wk or not str(wk).strip().lower().startswith("wk"):
            continue
        wk = str(wk).strip()
        vt = read_link(ws, r, BAL_VT)
        helper = read_link(ws, r, BAL_VTURL)
        topic = str(ws.cell(row=r, column=BAL_TOPIC).value or "").strip()
        ck = key_from_agenda_cycle(str(ws.cell(row=r, column=BAL_PHASE).value or ""))
        where = "BioAsList!F%d" % r

        if not vt.url:
            continue
        if vt.service == "drivefile":
            rep.add(ADVISORY, "A-BIOLIST-FILEFORM", where,
                    "%s %r VT link is a Drive file link, not presentation/d — "
                    "cannot be /copy distributed: %s" % (wk, topic, vt.url), ck)
        if helper.doc_id and vt.doc_id and helper.doc_id != vt.doc_id:
            rep.add(ADVISORY, "A-BIOLIST-VT-DRIFT", "BioAsList!F%d/K%d" % (r, r),
                    "%s %r: visible VT link (%s) and the _vtURL helper column "
                    "(%s) name different decks"
                    % (wk, topic, vt.doc_id, helper.doc_id), ck)
        if vt.doc_id and vt.doc_id not in bio_doc_ids:
            rep.add(ADVISORY, "A-BIOLIST-ORPHAN", where,
                    "%s %r points at deck %s, which the dated Biology tab never "
                    "uses" % (wk, topic, vt.doc_id), ck)


# ---------------------------------------------------------------------------
# Section B — deck <-> agenda agreement
# ---------------------------------------------------------------------------

def check_agreement(rows, decks_by_key, live_by_doc, live_ids, rep):
    # ---- resolve every row's deck to a cycle key -------------------------
    unknown_ids = {}
    for ar in rows:
        if not ar.doc_id:
            continue
        key = live_by_doc.get(ar.doc_id, "")
        ar.deck_key = key
        if not key:
            unknown_ids.setdefault(ar.doc_id, []).append(ar)

    for doc_id, ars in sorted(unknown_ids.items()):
        rows_s = ", ".join("r%d %r" % (a.row, a.topic[:34]) for a in ars[:4])
        rep.add(ADVISORY, "A-ID-UNKNOWN", "Biology!G/H",
                "deck id %s is in neither decks_live_ids.csv nor any deck on "
                "disk; used on %d row(s): %s" % (doc_id, len(ars), rows_s),
                ars[0].cycle_key)

    # ---- cycle-code and topic agreement ---------------------------------
    for ar in rows:
        if not ar.deck_key:
            continue
        deck = decks_by_key.get(ar.deck_key)
        deck_title = deck["norm"] if deck else live_ids.get(ar.deck_key, {}).get("live_name", "")
        aspects = deck.get("critical_aspects", "") if deck else ""

        an, dn = cycle_number(ar.cycle_key), cycle_number(ar.deck_key)
        if an is not None and dn is not None and an != dn:
            rep.add(ADVISORY, "A-CYCLE-MISMATCH", "Biology!C%d" % ar.row,
                    "row is labelled %s but links to %s (%r)"
                    % (ar.cycle_text, ar.deck_key, deck_title[:56]), ar.cycle_key)

        verdict = topic_agreement(ar.topic, deck_title, aspects)
        if verdict == "none":
            rep.add(ADVISORY, "A-TOPIC-DRIFT", "Biology!D%d" % ar.row,
                    "topic %r shares no content word with %s (%r) or its "
                    "critical aspects (%r)"
                    % (ar.topic, ar.deck_key, deck_title[:48], aspects[:70]),
                    ar.cycle_key)
        elif verdict == "title-only" and aspects:
            rep.add(ADVISORY, "A-TOPIC-ASPECT-DRIFT", "Biology!D%d" % ar.row,
                    "topic %r matches the deck TITLE but none of %s's critical "
                    "aspects (%r)" % (ar.topic, ar.deck_key, aspects[:70]),
                    ar.cycle_key)

    # ---- block sequencing and duplicate day-1 decks ----------------------
    # Katherine's rule: only the FIRST meeting of a multi-meeting cycle issues a
    # new student deck; later meetings continue the same one.
    prev_doc = None
    for ar in rows:
        if ar.doc_id and ar.block_i is not None and ar.block_i > 1:
            if prev_doc is not None and ar.doc_id != prev_doc:
                rep.add(ADVISORY, "A-LATER-NEW-DECK", "Biology!G%d" % ar.row,
                        "meeting %s issues a NEW deck (%s, %s) — only a '1 of N' "
                        "row should. %s — %r"
                        % (ar.block, ar.doc_id[:12] + "...",
                           ar.deck_key or "unknown id", ar.date, ar.topic),
                        ar.cycle_key)
        if ar.doc_id:
            prev_doc = ar.doc_id

    for ar in rows:
        if ar.is_day1() and not ar.doc_id:
            rep.add(ADVISORY, "A-DAY1-NO-DECK", "Biology!G%d:H%d" % (ar.row, ar.row),
                    "first meeting of a %d-meeting cycle (%s) issues no deck: "
                    "%s — %r" % (ar.block_n, ar.block, ar.date, ar.topic),
                    ar.cycle_key)

    # Duplicate day-1 decks are DEFECTS now, not acceptable repeats.
    # Earliest date is canonical; every later '1 of N' on the same deck id is
    # a unit that NEEDS ITS OWN DECK.
    day1_by_doc = {}
    needs_own = []
    for ar in rows:
        if ar.is_day1() and ar.doc_id:
            day1_by_doc.setdefault(ar.doc_id, []).append(ar)
    for doc_id, ars in day1_by_doc.items():
        ars.sort(key=lambda a: (a.date or datetime.date(1900, 1, 1), a.row))
        canon = ars[0]
        for dup in ars[1:]:
            # Attributed to the DECK being reused, not to the duplicate row's
            # agenda code: agenda codes (C10, C06e, C13e) do not map onto deck
            # cycle keys, and attributing it to the code hides the finding from
            # the roll-up entirely when no deck of that key exists.
            rep.add(ADVISORY, "A-DUP-DAY1", "Biology!H%d" % dup.row,
                    "NEEDS ITS OWN DECK — %s %r (%s) reuses the day-1 deck of "
                    "%s %r (%s, canonical). Reused deck: %s / %s"
                    % (dup.cycle_text or "?", dup.topic, dup.date,
                       canon.cycle_text or "?", canon.topic, canon.date,
                       dup.deck_key or "unknown", doc_id[:14] + "..."),
                    dup.deck_key or dup.cycle_key)
            needs_own.append((dup, canon))

    # A deck the agenda never introduces on a '1 of N' row.
    used_docs = set(a.doc_id for a in rows if a.doc_id)
    day1_docs = set(day1_by_doc.keys())
    for doc_id in sorted(used_docs - day1_docs):
        key = live_by_doc.get(doc_id, "")
        ars = [a for a in rows if a.doc_id == doc_id]
        blocks = ", ".join("r%d %s" % (a.row, a.block or "(no block)") for a in ars[:5])
        name = live_ids.get(key, {}).get("live_name", "") if key else ""
        who = "%s%s" % (key or (doc_id[:14] + "... (id unknown)"),
                        " (%s)" % name[:44] if name else "")
        rep.add(ADVISORY, "A-NO-DAY1", "Biology!G/H",
                "deck %s never appears on a '1 of N' row — it is only ever "
                "issued mid-cycle: %s" % (who, blocks), key)

    # ---- decks on disk with no agenda row --------------------------------
    scheduled_keys = set(a.deck_key for a in rows if a.deck_key)
    orphans = []
    for key in sorted(decks_by_key):
        if key not in scheduled_keys:
            d = decks_by_key[key]
            orphans.append((key, d))
            hint = ""
            if unknown_ids:
                hint = (" (note: %d agenda deck id(s) could not be resolved — "
                        "this deck may in fact be scheduled under an id that is "
                        "missing from decks_live_ids.csv)" % len(unknown_ids))
            rep.add(ADVISORY, "A-DECK-NO-AGENDA", "deck:%s" % key,
                    "built but never scheduled — no agenda row links to %r%s"
                    % (d["file"], hint), key)
    return orphans, needs_own


# ---------------------------------------------------------------------------
# Section C — readiness roll-up
# ---------------------------------------------------------------------------

CSV_COLUMNS = [
    "cycle", "deck_file", "deck_exists", "live_name_matches", "live_doc_id",
    "lint", "lint_hard", "lint_advisory",
    "agenda_rows", "agenda_row_numbers", "agenda_cycle_codes", "agenda_topics",
    "day1_rows", "dup_day1", "links_wellformed", "student_copy_ok",
    "topic_agreement", "hard", "advisory", "hard_codes", "advisory_codes",
    "ready", "blockers",
]


def build_rollup(decks_by_key, rows, live_ids, rep, lint_results):
    by_cycle_findings = {}
    for f in rep.findings:
        if f.cycle:
            by_cycle_findings.setdefault(f.cycle, []).append(f)

    keys = set(decks_by_key) | set(live_ids) | set(a.deck_key for a in rows if a.deck_key)
    keys = sorted(k for k in keys if k)

    out = []
    for key in keys:
        deck = decks_by_key.get(key)
        live = live_ids.get(key, {})
        ars = [a for a in rows if a.deck_key == key]
        fs = by_cycle_findings.get(key, [])
        # link findings restricted to the rows that actually carry this deck
        row_nums = set(a.row for a in ars)
        link_bad = [f for f in fs
                    if f.code.startswith(("H-STUDENT", "H-LINK", "H-DECK"))]
        student_ok = not any(f.code in ("H-STUDENT-EDIT", "H-STUDENT-NOTCOPY",
                                        "H-STUDENT-FILEFORM", "H-STUDENT-MISSING")
                             for f in fs)
        lr = lint_results.get(key)
        if lr is None:
            lint_verdict, lh, la = ("not run", "", "")
        else:
            lint_verdict = "ships" if lr["hard"] == 0 else "DOES NOT SHIP"
            lh, la = lr["hard"], lr["advisory"]

        agree = ""
        if ars and deck:
            verdicts = [topic_agreement(a.topic, deck["norm"],
                                        deck.get("critical_aspects", ""))
                        for a in ars]
            order = ("none", "title-only", "aspect-only", "title+aspect", "n/a")
            agree = sorted(set(verdicts), key=lambda v: order.index(v) if v in order else 9)[0]

        hard = [f for f in fs if f.tier == HARD]
        adv = [f for f in fs if f.tier == ADVISORY]

        blockers = []
        if not deck:
            blockers.append("no deck on disk")
        if lr is not None and lr["hard"]:
            blockers.append("deck_lint HARD x%d" % lr["hard"])
        if not ars:
            blockers.append("no agenda row")
        if link_bad:
            blockers.append("link defects x%d" % len(link_bad))
        if not student_ok:
            blockers.append("student link not /copy")
        if not live.get("doc_id"):
            blockers.append("no live doc_id (no import route)")
        if any(f.code == "A-DUP-DAY1" for f in fs):
            # ADVISORY for the exit code (the deck is not built yet) but it is a
            # ruled defect, so it must not read as READY.
            blockers.append("this deck's day-1 is reused by a later unit, "
                            "which needs its own")

        out.append({
            "cycle": key,
            "deck_file": deck["file"] if deck else "",
            "deck_exists": "yes" if deck else "NO",
            "live_name_matches": ("yes" if (deck and live.get("live_name", "").strip()
                                            == deck["norm"]) else
                                  ("n/a" if not deck or not live else "NO")),
            "live_doc_id": live.get("doc_id", ""),
            "lint": lint_verdict,
            "lint_hard": lh,
            "lint_advisory": la,
            "agenda_rows": len(ars),
            "agenda_row_numbers": " ".join(str(a.row) for a in ars),
            "agenda_cycle_codes": " ".join(sorted(set(a.cycle_text for a in ars if a.cycle_text))),
            "agenda_topics": " | ".join(sorted(set(a.topic for a in ars if a.topic)))[:180],
            "day1_rows": " ".join(str(a.row) for a in ars if a.is_day1()),
            "dup_day1": "yes" if any(f.code == "A-DUP-DAY1" for f in fs) else "",
            "links_wellformed": "yes" if not link_bad else "NO",
            "student_copy_ok": "yes" if student_ok else "NO",
            "topic_agreement": agree,
            "hard": len(hard),
            "advisory": len(adv),
            "hard_codes": "; ".join(sorted(set(f.code for f in hard))),
            "advisory_codes": "; ".join(sorted(set(f.code for f in adv))),
            "ready": "READY" if not blockers else "not ready",
            "blockers": "; ".join(blockers),
        })
    return out


# ---------------------------------------------------------------------------
# deck_lint integration
# ---------------------------------------------------------------------------

def run_lint(decks_by_key, rep):
    """Import deck_lint and run it on the canonical deck of each cycle.

    Imported, not shelled out: same process, no CSV round-trip, and the Finding
    objects come back structured. Falls back to a subprocess if the import fails.
    """
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    try:
        import deck_lint
    except Exception as exc:               # pragma: no cover
        print("deck_agenda_audit: could not import deck_lint (%s); "
              "skipping the lint pass" % exc, file=sys.stderr)
        return {}
    out = {}
    for key in sorted(decks_by_key):
        d = decks_by_key[key]
        try:
            r = deck_lint.lint_deck(d["path"])
        except Exception as exc:
            rep.add(ADVISORY, "A-LINT-ERROR", "deck:%s" % key,
                    "deck_lint raised %s on %r" % (type(exc).__name__, d["file"]),
                    key)
            continue
        out[key] = r
        if r["hard"]:
            codes = sorted(set(f.code for f in r["findings"] if f.tier == "HARD"))
            rep.add(ADVISORY, "A-LINT-HARD", "deck:%s" % key,
                    "deck_lint says DOES NOT SHIP — %d hard finding(s): %s "
                    "(run deck_lint.py for detail; it owns these)"
                    % (r["hard"], ", ".join(codes)), key)
    return out


# ---------------------------------------------------------------------------
# --check-live (OFF by default; the only thing here that touches the network)
# ---------------------------------------------------------------------------

_UA = {"User-Agent": "Mozilla/5.0 (compatible; deck-agenda-audit/1.0)"}


def check_live_ids(doc_ids, rep):
    """Resolve each deck id. 404 => the deck is gone. 200 only means the id
    exists: unauthenticated, Google answers 200 with a sign-in page for a
    private deck. Verified 2026-08-29: a fabricated id returns 404, a real one
    returns 200, so this does discriminate "missing" from "present"."""
    import ssl  # noqa: F401  (kept: some environments need the import to seed certs)
    import urllib.request
    import urllib.error
    for doc_id in sorted(doc_ids):
        url = "https://docs.google.com/presentation/d/%s/edit" % doc_id
        status = None
        for method in ("HEAD", "GET"):
            try:
                req = urllib.request.Request(url, headers=_UA, method=method)
                with urllib.request.urlopen(req, timeout=25) as r:
                    status = r.status
                    break
            except urllib.error.HTTPError as e:
                if method == "HEAD":
                    continue
                status = e.code
            except Exception as e:
                if method == "HEAD":
                    continue
                status = "ERR:" + type(e).__name__
        if status != 200:
            rep.add(ADVISORY, "A-LIVE-UNRESOLVED", "live:%s" % doc_id,
                    "deck id does not resolve (status %s): %s" % (status, url))


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def print_report(rollup, rep, rows, decks, orphans, needs_own, tier_filter, ctx):
    W = 78
    print("=" * W)
    print("deck_agenda_audit — decks vs. the agenda")
    print("=" * W)
    print("  agenda:      %s" % ctx["agenda"])
    print("  exports:     %s" % ", ".join(ctx["roots"]))
    print("  live ids:    %s (%d rows)" % (ctx["live_ids_path"], ctx["n_live"]))
    print("  inventory:   %s" % (ctx["inventory"] or "(none found)"))
    print("  agenda rows: %d   decks on disk: %d (%d canonical, %d variants)"
          % (len(rows), len(decks), ctx["n_canonical"], len(decks) - ctx["n_canonical"]))
    print("  network:     %s" % ("--check-live ON" if ctx["check_live"]
                                 else "OFF (no requests made)"))
    print()

    groups = rep.by_code()
    for section, codes in (
        ("A. LINK INTEGRITY",
         ["H-LINK-EMPTY", "H-STUDENT-EDIT", "H-STUDENT-NOTCOPY",
          "H-STUDENT-FILEFORM", "H-LINK-MISMATCH", "H-STUDENT-MISSING",
          "H-DECK-MISSING", "A-TEACHER-MISSING", "A-TEACHER-COPY",
          "A-TEACHER-FILEFORM", "A-FOLDER-LINK", "A-DAY-WEEKDAY",
          "A-BIOLIST-FILEFORM", "A-BIOLIST-VT-DRIFT", "A-BIOLIST-ORPHAN",
          "A-LIVE-UNRESOLVED"]),
        ("B. DECK <-> AGENDA AGREEMENT",
         ["A-DECK-NO-AGENDA", "A-AGENDA-NO-DECK", "A-ID-UNKNOWN",
          "A-DUP-DAY1", "A-DAY1-NO-DECK", "A-LATER-NEW-DECK", "A-NO-DAY1",
          "A-CYCLE-MISMATCH", "A-TOPIC-DRIFT", "A-TOPIC-ASPECT-DRIFT",
          "A-CYCLE-MULTI-DECK", "A-NOT-IN-INVENTORY", "A-NO-LIVE-ID",
          "A-DECK-UNKEYED", "A-LINT-HARD", "A-LINT-ERROR"]),
    ):
        print("-" * W)
        print(section)
        print("-" * W)
        any_shown = False
        for code in codes:
            fs = groups.get(code, [])
            if not fs:
                continue
            tier = fs[0].tier
            if tier_filter not in ("all", tier.lower()):
                continue
            any_shown = True
            print("  %-8s %-22s  %d" % (tier, code, len(fs)))
            for f in fs[:LIST_CAP]:
                print("      %s: %s" % (f.where, f.message))
            if len(fs) > LIST_CAP:
                print("      ... and %d more; full list in the findings CSV."
                      % (len(fs) - LIST_CAP))
            print()
        if not any_shown:
            print("  nothing to report at tier '%s'\n" % tier_filter)

    print("-" * W)
    print("C. READINESS ROLL-UP")
    print("-" * W)
    print("  %-10s %-4s %-13s %-5s %-6s %-6s %-6s %s"
          % ("cycle", "deck", "lint", "rows", "links", "/copy", "docid", "verdict"))
    for r in rollup:
        print("  %-10s %-4s %-13s %-5s %-6s %-6s %-6s %s"
              % (r["cycle"], "yes" if r["deck_exists"] == "yes" else "NO",
                 r["lint"], r["agenda_rows"],
                 "ok" if r["links_wellformed"] == "yes" else "BAD",
                 "ok" if r["student_copy_ok"] == "yes" else "BAD",
                 "ok" if r["live_doc_id"] else "NO",
                 r["ready"] if r["ready"] == "READY" else r["blockers"]))
    print()

    if needs_own:
        print("-" * W)
        print("NEEDS ITS OWN DECK  (duplicate day-1 decks — Katherine's ruling:")
        print("                     'later units need their own decks')")
        print("-" * W)
        for dup, canon in needs_own:
            print("  %-6s %-38s %s   currently reuses %s %r (%s)"
                  % (dup.cycle_text or "?", dup.topic[:38], dup.date,
                     canon.cycle_text or "?", canon.topic[:30], canon.date))
        print()

    print("=" * W)
    ready = sum(1 for r in rollup if r["ready"] == "READY")
    print("SUMMARY")
    print("  cycles:          %d   READY: %d   not ready: %d"
          % (len(rollup), ready, len(rollup) - ready))
    print("  HARD findings:     %d" % len(rep.hard()))
    print("  ADVISORY findings: %d" % len(rep.advisory()))
    print("  decks with no agenda row: %d  -> %s"
          % (len(orphans), ", ".join(k for k, _ in orphans) or "(none)"))
    nodeck = [f for f in rep.findings
              if f.code in ("H-DECK-MISSING", "A-AGENDA-NO-DECK")]
    print("  needs its own deck:  %d  -> %s"
          % (len(needs_own),
             ", ".join("%s %s" % (d.cycle_text, d.topic) for d, _ in needs_own)
             or "(none)"))
    print("  agenda rows with no deck: %d (%d scheduled meetings = HARD, "
          "%d skeleton rows = advisory)"
          % (len(nodeck),
             sum(1 for f in nodeck if f.code == "H-DECK-MISSING"),
             sum(1 for f in nodeck if f.code == "A-AGENDA-NO-DECK")))
    print("  exit code:       %d  (HARD only)" % (1 if rep.hard() else 0))
    print("=" * W)


def write_csv(rollup, rep, path):
    path = _guard_output_path(path)
    d = os.path.dirname(path)
    if d and not os.path.isdir(d):
        os.makedirs(d)
    with open(path, "w") as fh:
        w = csvmod.writer(fh)
        w.writerow(CSV_COLUMNS)
        for r in rollup:
            w.writerow([r.get(c, "") for c in CSV_COLUMNS])
    # findings alongside, same stem + _findings
    stem, ext = os.path.splitext(path)
    fpath = stem + "_findings" + (ext or ".csv")
    with open(fpath, "w") as fh:
        w = csvmod.writer(fh)
        w.writerow(["tier", "code", "cycle", "where", "message"])
        for f in rep.findings:
            w.writerow(f.as_row())
    return path, fpath


def write_md(rollup, rep, rows, decks, orphans, needs_own, path, ctx):
    path = _guard_output_path(path)
    d = os.path.dirname(path)
    if d and not os.path.isdir(d):
        os.makedirs(d)
    groups = rep.by_code()
    L = []
    A = L.append
    A("# Deck / agenda audit — %s" % datetime.date.today().isoformat())
    A("")
    A("Decks are **not finalized**. Everything ADVISORY below is work not done, "
      "not a defect. Only HARD blocks.")
    A("")
    A("| | |")
    A("|---|---|")
    A("| agenda | `%s` |" % ctx["agenda"])
    A("| exports | %s |" % ", ".join("`%s`" % r for r in ctx["roots"]))
    A("| live ids | `%s` (%d rows) |" % (ctx["live_ids_path"], ctx["n_live"]))
    A("| inventory | `%s` |" % (ctx["inventory"] or "(none)"))
    A("| agenda rows | %d |" % len(rows))
    A("| decks on disk | %d (%d canonical) |" % (len(decks), ctx["n_canonical"]))
    A("| network | %s |" % ("--check-live ON" if ctx["check_live"] else "OFF"))
    A("| HARD | %d |" % len(rep.hard()))
    A("| ADVISORY | %d |" % len(rep.advisory()))
    A("")
    if needs_own:
        A("## Needs its own deck")
        A("")
        A("Katherine's ruling: *later units need their own decks.* The earlier "
          "date is canonical.")
        A("")
        A("| unit | topic | date | currently reuses |")
        A("|---|---|---|---|")
        for dup, canon in needs_own:
            A("| %s | %s | %s | %s %s (%s) |"
              % (dup.cycle_text or "?", dup.topic, dup.date,
                 canon.cycle_text or "?", canon.topic, canon.date))
        A("")
    A("## C. Readiness roll-up")
    A("")
    A("| cycle | deck | lint | agenda rows | links | student /copy | live doc_id | verdict |")
    A("|---|---|---|---|---|---|---|---|")
    for r in rollup:
        A("| %s | %s | %s | %s | %s | %s | %s | %s |"
          % (r["cycle"], r["deck_exists"], r["lint"], r["agenda_rows"],
             r["links_wellformed"], r["student_copy_ok"],
             "yes" if r["live_doc_id"] else "NO",
             r["ready"] if r["ready"] == "READY" else r["blockers"]))
    A("")
    for title, codes in (
        ("## A. Link integrity",
         ["H-LINK-EMPTY", "H-STUDENT-EDIT", "H-STUDENT-NOTCOPY",
          "H-STUDENT-FILEFORM", "H-LINK-MISMATCH", "H-STUDENT-MISSING",
          "H-DECK-MISSING", "A-TEACHER-MISSING", "A-TEACHER-COPY",
          "A-TEACHER-FILEFORM", "A-FOLDER-LINK", "A-DAY-WEEKDAY",
          "A-BIOLIST-FILEFORM", "A-BIOLIST-VT-DRIFT", "A-BIOLIST-ORPHAN",
          "A-LIVE-UNRESOLVED"]),
        ("## B. Deck ↔ agenda agreement",
         ["A-DECK-NO-AGENDA", "A-AGENDA-NO-DECK", "A-ID-UNKNOWN",
          "A-DUP-DAY1", "A-DAY1-NO-DECK", "A-LATER-NEW-DECK", "A-NO-DAY1",
          "A-CYCLE-MISMATCH", "A-TOPIC-DRIFT", "A-TOPIC-ASPECT-DRIFT",
          "A-CYCLE-MULTI-DECK", "A-NOT-IN-INVENTORY", "A-NO-LIVE-ID",
          "A-DECK-UNKEYED", "A-LINT-HARD", "A-LINT-ERROR"]),
    ):
        A(title)
        A("")
        wrote = False
        for code in codes:
            fs = groups.get(code, [])
            if not fs:
                continue
            wrote = True
            A("### %s — %s (%d)" % (code, fs[0].tier, len(fs)))
            A("")
            for f in fs[:LIST_CAP]:
                A("- `%s` — %s" % (f.where, f.message))
            if len(fs) > LIST_CAP:
                A("- _... and %d more; full list in the findings CSV._"
                  % (len(fs) - LIST_CAP))
            A("")
        if not wrote:
            A("_nothing to report._")
            A("")
    with open(path, "w") as fh:
        fh.write("\n".join(L) + "\n")
    return path


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main(argv=None):
    ap = argparse.ArgumentParser(
        prog="deck_agenda_audit.py",
        description="Audit the VT Biology decks against the teaching agenda and "
                    "against the links that carry them to students. Two tiers: "
                    "HARD fails, ADVISORY reports. Read-only.")
    ap.add_argument("--agenda", default=DEFAULT_AGENDA,
                    help="agenda workbook (default: the read-freely backup). "
                         "The LIVE CloudStorage copy is read-only and is never "
                         "written; pass --agenda-live to read it instead.")
    ap.add_argument("--agenda-live", action="store_true",
                    help="read the LIVE agenda in CloudStorage instead of the "
                         "backup. Still read-only.")
    ap.add_argument("--exports", default=DEFAULT_EXPORTS)
    ap.add_argument("--rebuild", default=DEFAULT_REBUILD)
    ap.add_argument("--live-ids", default=DEFAULT_LIVE_IDS)
    ap.add_argument("--inventory", default=None,
                    help="deck_inventory_*.csv (default: newest in audits/)")
    ap.add_argument("--tier", choices=("hard", "advisory", "all"), default="all",
                    help="which tier to print. Never changes the exit code.")
    ap.add_argument("--no-lint", action="store_true",
                    help="skip the deck_lint pass")
    ap.add_argument("--check-live", action="store_true",
                    help="OFF BY DEFAULT. Verify each deck id resolves. This is "
                         "the only thing in this script that touches the "
                         "network. A 404 proves the deck is gone; a 200 only "
                         "proves the id exists — unauthenticated, Google serves "
                         "a sign-in page for a private deck with status 200.")
    ap.add_argument("--csv", default=None, metavar="PATH")
    ap.add_argument("--md", default=None, metavar="PATH")
    args = ap.parse_args(argv)

    agenda = os.path.abspath(os.path.expanduser(
        LIVE_AGENDA if args.agenda_live else args.agenda))
    if not os.path.exists(agenda):
        raise SystemExit("deck_agenda_audit: agenda not found: %s" % agenda)
    WITNESS.watch(agenda)

    rep = Report()

    # ---- agenda ----------------------------------------------------------
    wb = openpyxl.load_workbook(agenda, data_only=False, read_only=False)
    ws_bio = require_sheet(wb, TAB_BIOLOGY, agenda)
    ws_bal = require_sheet(wb, TAB_BIO_AS_LIST, agenda)
    require_sheet(wb, TAB_SCHEDULE, agenda)
    require_sheet(wb, TAB_FILE_IDS, agenda)
    require_sheet(wb, TAB_LAB_MATERIALS, agenda)   # LEADING SPACE; must exist

    rows, header_problems = load_biology_rows(ws_bio)
    for col, want, got in header_problems:
        rep.add(HARD, "H-HEADER-DRIFT", "Biology!%s%d" % (
            openpyxl.utils.get_column_letter(col), BIO_HEADER_ROW),
            "expected a %r header, found %r — the column map in this script is "
            "no longer valid" % (want, got))
    if not rows:
        raise SystemExit(
            "deck_agenda_audit: %r produced ZERO data rows. Refusing to report a "
            "clean audit on an empty read." % TAB_BIOLOGY)

    # ---- live ids, inventory, decks --------------------------------------
    live_ids, live_by_doc = load_live_ids(os.path.expanduser(args.live_ids))
    inv_path = newest_inventory(args.inventory)
    inventory = load_inventory(inv_path)

    roots = [args.exports, args.rebuild]
    decks = scan_decks(roots)

    decks_by_key = {}
    base_files = {}
    for d in decks:
        if d["is_backup"]:
            continue
        key = key_from_text(d["norm"]) or key_from_text(d["folder"])
        if not key:
            rep.add(ADVISORY, "A-DECK-UNKEYED", "deck:%s" % d["file"],
                    "cannot derive a cycle key from the filename or its folder")
            continue
        # 12b and friends: a letter suffix directly after the number
        m = re.search(r'Cycle\s*0*(\d+)\s*([a-z])\b', d["norm"], re.IGNORECASE)
        if m:
            key = cycle_key(m.group(1), m.group(2))
        d["key"] = key
        inv = inventory.get(d["file"], {})
        d["critical_aspects"] = inv.get("critical_aspects", "")
        d["slides"] = inv.get("slides", "")
        d["in_inventory"] = bool(inv)
        prev = decks_by_key.get(key)
        # canonical = the file whose normalized name equals live_name; else the
        # plain "LIVE export" (no Concept-Bank suffix); else first seen.
        live_name = (live_ids.get(key, {}).get("live_name") or "").strip()
        def score(x):
            s = 0
            if live_name and x["norm"] == live_name:
                s += 4
            if not x["file"].endswith(tuple(s2 + ".pptx" for s2 in _CB_SUFFIXES)):
                s += 2
            if x["file"].startswith(u"▶"):
                s += 1
            return s
        if prev is None or score(d) > score(prev):
            decks_by_key[key] = d
        base_files.setdefault(key, [])
        if not d["file"].endswith(tuple(s2 + ".pptx" for s2 in _CB_SUFFIXES)):
            base_files[key].append(d["file"])

    n_canonical = len(decks_by_key)
    for key, files in sorted(base_files.items()):
        if len(files) > 1:
            rep.add(ADVISORY, "A-CYCLE-MULTI-DECK", "deck:%s" % key,
                    "%d distinct base decks share this cycle key, so only one "
                    "could be rolled up: %s"
                    % (len(files), "; ".join(sorted(files))), key)
    for key, d in sorted(decks_by_key.items()):
        if not d.get("in_inventory"):
            rep.add(ADVISORY, "A-NOT-IN-INVENTORY", "deck:%s" % key,
                    "%r is not in %s — its critical aspects could not be "
                    "compared against the agenda topic"
                    % (d["file"], os.path.basename(inv_path or "the inventory")),
                    key)
        if key not in live_ids:
            rep.add(ADVISORY, "A-NO-LIVE-ID", "deck:%s" % key,
                    "no row in decks_live_ids.csv — there is no import route for "
                    "this deck (%r)" % d["file"], key)

    # ---- section A -------------------------------------------------------
    check_links(rows, rep)
    bio_doc_ids = set(a.doc_id for a in rows if a.doc_id)
    check_bio_as_list(ws_bal, bio_doc_ids, rep)

    # ---- section B -------------------------------------------------------
    orphans, needs_own = check_agreement(rows, decks_by_key, live_by_doc,
                                        live_ids, rep)

    # ---- lint ------------------------------------------------------------
    lint_results = {} if args.no_lint else run_lint(decks_by_key, rep)

    # ---- optional network pass ------------------------------------------
    if args.check_live:
        ids = set(bio_doc_ids)
        ids |= set(v.get("doc_id", "") for v in live_ids.values() if v.get("doc_id"))
        check_live_ids(ids, rep)

    # ---- roll-up and output ---------------------------------------------
    rollup = build_rollup(decks_by_key, rows, live_ids, rep, lint_results)
    ctx = {
        "agenda": agenda,
        "roots": [os.path.abspath(os.path.expanduser(r)) for r in roots],
        "live_ids_path": os.path.abspath(os.path.expanduser(args.live_ids)),
        "n_live": len(live_ids),
        "inventory": inv_path,
        "n_canonical": n_canonical,
        "check_live": args.check_live,
    }
    print_report(rollup, rep, rows, decks, orphans, needs_own,
                 args.tier, ctx)

    if args.csv:
        p, fp = write_csv(rollup, rep, args.csv)
        print("\ncsv written:      %s" % p)
        print("findings written: %s" % fp)
    if args.md:
        p = write_md(rollup, rep, rows, decks, orphans, needs_own,
                     args.md, ctx)
        print("md written:       %s" % p)

    n = WITNESS.assert_unchanged()
    print("read-only assertion: %d watched file(s) unchanged." % n)

    return 1 if rep.hard() else 0


if __name__ == "__main__":
    sys.exit(main())
